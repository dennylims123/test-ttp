# Convert "list ID Desa.xlsx" to the new villages.json format with msd_status
import openpyxl
import json
from collections import Counter

wb = openpyxl.load_workbook('/home/z/my-project/upload/list ID Desa.xlsx', data_only=True)
ws = wb['List ID Desa']
print(f"Total rows: {ws.max_row}")

villages = []
msd_counts = Counter()

for r in range(2, ws.max_row + 1):
    village = ws.cell(row=r, column=1).value  # Column A: Village
    subdistrict = ws.cell(row=r, column=2).value  # Column B: SubDistrict
    district = ws.cell(row=r, column=3).value  # Column C: District
    province = ws.cell(row=r, column=4).value  # Column D: Province
    full_str = ws.cell(row=r, column=8).value  # Column H: "Village, SubDist, District, Province"
    msd = ws.cell(row=r, column=9).value  # Column I: MSD / Non-MSD / #N/A

    if not village:
        continue

    # Use full_str if available, otherwise build it
    if full_str:
        full = str(full_str).strip()
    else:
        full = ', '.join([str(x) for x in [village, subdistrict, district, province] if x])

    # Normalize MSD status
    msd_status = str(msd).strip() if msd else '#N/A'
    if msd_status not in ('MSD', 'Non-MSD', '#N/A'):
        msd_status = '#N/A'
    msd_counts[msd_status] += 1

    villages.append({
        'id': r - 1,  # 1-indexed
        'desa': str(village).strip(),
        'full': full,
        'msd_status': msd_status,
    })

print(f"\nConverted {len(villages)} villages")
print("MSD distribution:")
for k, v in msd_counts.most_common():
    print(f"  {k}: {v}")

# Save to public/data/villages.json
with open('/home/z/my-project/public/data/villages.json', 'w', encoding='utf-8') as f:
    json.dump(villages, f, ensure_ascii=False)
print(f"\nSaved to public/data/villages.json ({len(json.dumps(villages)) // 1024} KB)")
